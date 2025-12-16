"""
자동 학급 편성 프로그램
5학년 152명의 학생을 7개 반으로 균등하게 배정하는 시스템
"""

import pandas as pd
import numpy as np
from dataclasses import dataclass, field
from typing import List, Dict, Set, Tuple, Optional
import random
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.comments import Comment
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox


@dataclass
class Student:
    """학생 정보를 담는 데이터 클래스"""
    학년: int
    원반: int  # 원래 5학년 반
    원번호: int  # 원래 번호
    이름: str
    성별: str
    점수: float
    특수반: bool
    전출: bool
    난이도: float
    비고: str

    # 배정 관련 필드
    assigned_class: Optional[int] = None  # 배정된 6학년 반 (1-7)
    locked: bool = False  # 배정 후 변경 불가 플래그
    rank: Optional[int] = None  # 성별 내 등수 (남학생이면 남학생 중, 여학생이면 여학생 중)

    def __post_init__(self):
        # NaN 처리
        if pd.isna(self.특수반):
            self.특수반 = False
        else:
            self.특수반 = bool(self.특수반)

        if pd.isna(self.전출):
            self.전출 = False
        else:
            self.전출 = bool(self.전출)

        if pd.isna(self.난이도):
            self.난이도 = 0.0

        if pd.isna(self.비고):
            self.비고 = ""

    def effective_count(self) -> int:
        """유효 인원 계산: 특수반=3명, 전출생=0명, 일반=1명"""
        if self.전출:
            return 0
        return 3 if self.특수반 else 1


class ClassAssigner:
    """학급 편성 시스템"""

    def __init__(self, student_file: str, rules_file: str, target_class_count: int = 7):
        self.student_file = student_file
        self.rules_file = rules_file
        self.target_class_count = target_class_count
        self.students: List[Student] = []
        self.classes: Dict[int, List[Student]] = {i: [] for i in range(1, self.target_class_count + 1)}

        # 규칙
        self.separation_rules: Dict[str, Set[str]] = defaultdict(set)  # 분반 규칙
        self.separation_pairs: List[Tuple[str, str]] = []  # 분반 쌍 (색상 구분용)
        self.together_groups: List[Set[str]] = []  # 합반 규칙

        print("=" * 70)
        print("🎓 자동 학급 편성 프로그램 시작")
        print("=" * 70)

    def load_students(self):
        """모든 시트에서 학생 데이터 로드"""
        print("\n📚 Step 0: 학생 데이터 로드 중...")

        all_students = []
        try:
            xl = pd.ExcelFile(self.student_file)
            sheet_names = xl.sheet_names
            print(f"   ℹ️  감지된 시트: {sheet_names}")
        except Exception as e:
            print(f"   ❌ 파일 읽기 오류: {e}")
            raise

        for sheet_name in sheet_names:
            try:
                # 숫자나 '-'가 포함된 시트만 반 정보로 간주 (필요 시 로직 강화 가능)
                # 현재는 모든 시트를 시도하되, 데이터 구조가 안 맞으면 스킵하는 방식이 안전할 수 있음
                df = pd.read_excel(self.student_file, sheet_name=sheet_name)

                # 필수 컬럼 확인
                required_cols = ['학년', '반', '번호', '이름']
                if not all(col in df.columns for col in required_cols):
                    print(f"   ⚠️  Skipping sheet '{sheet_name}': 필수 컬럼 누락")
                    continue


            except Exception as e:
                print(f"   ⚠️  Error reading sheet '{sheet_name}': {e}")
                continue

            for _, row in df.iterrows():
                student = Student(
                    학년=int(row['학년']),
                    원반=int(row['반']),
                    원번호=int(row['번호']),
                    이름=str(row['이름']),
                    성별=str(row['성별']),
                    점수=float(row['점수']),
                    특수반=row['특수반'],
                    전출=row['전출'],
                    난이도=row['난이도'],
                    비고=row['비고']
                )
                all_students.append(student)

        self.students = all_students

        # 성별별 등수 계산
        self._calculate_ranks()

        print(f"   ✅ 총 {len(self.students)}명의 학생 데이터 로드 완료")
        print(f"   - 남학생: {sum(1 for s in self.students if s.성별 == '남')}명")
        print(f"   - 여학생: {sum(1 for s in self.students if s.성별 == '여')}명")
        print(f"   - 특수반: {sum(1 for s in self.students if s.특수반)}명")
        print(f"   - 전출생: {sum(1 for s in self.students if s.전출)}명")

    def load_from_result(self, result_file: str):
        """
        배정 결과 파일(03 배정 결과.xlsx)을 로드

        Args:
            result_file: 배정 결과 Excel 파일 경로

        Returns:
            None (self.students, self.classes를 채움)
        """
        print("\n📂 배정 결과 파일 로드 중...")

        try:
            xl = pd.ExcelFile(result_file)
            sheet_names = xl.sheet_names

            # '요약', '규칙' 시트 제외
            exclude_sheets = ['요약', '규칙']
            class_sheets = [s for s in sheet_names if s not in exclude_sheets]

            if not class_sheets:
                raise ValueError("배정 결과 파일에 반 시트가 없습니다.")

            print(f"   ℹ️  감지된 반: {class_sheets}")

            # target_class_count 자동 설정
            self.target_class_count = len(class_sheets)
            self.classes = {i: [] for i in range(1, self.target_class_count + 1)}

            all_students = []

            # 각 반 시트 읽기
            for sheet_name in class_sheets:
                # 시트 이름에서 반 번호 추출 (예: '6-1' → 1)
                class_num = int(sheet_name.split('-')[1])

                df = pd.read_excel(result_file, sheet_name=sheet_name)

                for _, row in df.iterrows():
                    # Student 객체 생성
                    student = Student(
                        학년=int(row['원학년']),      # 원학년 사용 (5)
                        원반=int(row['원반']),
                        원번호=int(row['원번호']),
                        이름=str(row['이름']),
                        성별=str(row['성별']),
                        점수=float(row['점수']),
                        특수반=row['특수반'],
                        전출=row['전출'],
                        난이도=row['난이도'],
                        비고=row['비고'],
                        assigned_class=class_num      # 이미 배정됨
                    )

                    all_students.append(student)
                    self.classes[class_num].append(student)

            self.students = all_students

            # 성별별 등수 계산
            self._calculate_ranks()

            print(f"   ✅ 총 {len(self.students)}명 로드 완료")
            print(f"   - {self.target_class_count}개 반")
            for i in range(1, self.target_class_count + 1):
                print(f"     {i}반: {len(self.classes[i])}명")

            # 규칙 로드 (규칙 시트가 있으면)
            print()
            if '규칙' in sheet_names:
                print("📋 Step 1: 분반/합반 규칙 로드 중...")
                self._load_rules_from_sheet(result_file)
            else:
                print("   ⚠️  경고: 규칙 시트가 없습니다 (이전 버전 파일)")

        except Exception as e:
            print(f"   ❌ 파일 읽기 오류: {e}")
            raise

    @staticmethod
    def is_result_file(file_path: str) -> bool:
        """
        파일이 배정 결과 파일인지 확인

        Args:
            file_path: Excel 파일 경로

        Returns:
            True if 배정 결과 파일, False otherwise
        """
        try:
            xl = pd.ExcelFile(file_path)

            # 배정 결과 파일 특징:
            # 1. '요약' 시트 존재
            # 2. '6-1', '6-2' 같은 형식의 시트 존재
            # 3. (선택) '규칙' 시트 존재 가능
            if '요약' in xl.sheet_names:
                # 형식: 숫자-숫자 (예: '6-1')
                import re
                pattern = r'^\d+-\d+$'
                # '요약', '규칙' 제외한 반 시트만 확인
                class_sheets = [s for s in xl.sheet_names
                              if re.match(pattern, s)]
                return len(class_sheets) > 0

            return False
        except:
            return False

    def _calculate_ranks(self):
        """성별별 등수 계산"""
        # 남학생 등수 부여
        males = sorted([s for s in self.students if s.성별 == '남'],
                      key=lambda s: s.점수, reverse=True)
        for rank, student in enumerate(males, 1):
            student.rank = rank

        # 여학생 등수 부여
        females = sorted([s for s in self.students if s.성별 == '여'],
                        key=lambda s: s.점수, reverse=True)
        for rank, student in enumerate(females, 1):
            student.rank = rank

    def load_rules(self):
        """분반/합반 규칙 로드 및 검증"""
        print("\n📋 Step 1: 분반/합반 규칙 로드 중...")

        df = pd.read_excel(self.rules_file, sheet_name='Sheet1')

        # 분반 규칙 파싱 (첫 5개 열)
        separation_count = 0
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 행 스킵
                continue

            student1_class = row['분반해야하는 학생']
            student1_name = row['Unnamed: 1']
            student2_class = row['Unnamed: 3']
            student2_name = row['Unnamed: 4']

            if pd.notna(student1_name) and pd.notna(student2_name):
                self.separation_rules[student1_name].add(student2_name)
                self.separation_rules[student2_name].add(student1_name)
                self.separation_pairs.append((student1_name, student2_name))  # 쌍 저장
                separation_count += 1

        # 합반 규칙 파싱 (마지막 5개 열)
        together_count = 0
        current_group = set()
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 행 스킵
                continue

            student1_name = row['Unnamed: 7']  # 왼쪽 이름
            student2_name = row['Unnamed: 10']  # 오른쪽 이름

            # 왼쪽 또는 오른쪽에 학생 이름이 있으면 그룹에 추가
            if pd.notna(student1_name) or pd.notna(student2_name):
                if pd.notna(student1_name):
                    current_group.add(student1_name)
                if pd.notna(student2_name):
                    current_group.add(student2_name)
            else:
                # 둘 다 비어있으면 그룹 종료
                if current_group:
                    self.together_groups.append(current_group)
                    together_count += len(current_group)
                    current_group = set()

        if current_group:
            self.together_groups.append(current_group)
            together_count += len(current_group)

        print(f"   ✅ 분반 규칙: {separation_count}쌍")
        print(f"   ✅ 합반 규칙: {len(self.together_groups)}그룹 ({together_count}명)")

        # 규칙 충돌 검증
        self._validate_rules()

    def _validate_rules(self):
        """규칙 간 논리적 모순 검증"""
        print("   🔍 규칙 충돌 검증 중...")

        conflicts = []

        # 학생 명단에서 이름별 인원수 계산 (동명이인 확인용)
        name_counts = Counter(s.이름 for s in self.students)

        # 1. 합반 그룹 내부에서 분반 규칙 검사
        for group in self.together_groups:
            for name1 in group:
                for name2 in group:
                    if name1 != name2 and name2 in self.separation_rules.get(name1, set()):
                        conflicts.append(f"❌ 충돌: {name1}와 {name2}는 합반해야 하지만 동시에 분반해야 함")

        # 2. 합반 그룹 내부에서 동명이인 검사 (학생 명단 기준)
        for group in self.together_groups:
            for name in group:
                if name_counts[name] > 1:
                    conflicts.append(f"❌ 충돌: '{name}' 학생은 동명이인({name_counts[name]}명)이므로 합반 규칙에 포함될 수 없음")

        if conflicts:
            print("\n" + "=" * 70)
            print("⚠️  규칙 충돌 발견!")
            print("=" * 70)
            for conflict in conflicts:
                print(conflict)
            print("\n💡 해결 방법: '02 분반 합반할 학생 규칙.xlsx' 파일을 수정해주세요.")
            raise ValueError("규칙 충돌이 발견되었습니다. 위의 충돌을 해결한 후 다시 실행해주세요.")

        print("   ✅ 규칙 충돌 없음 - 모든 규칙이 논리적으로 일관됨")

    def _load_rules_from_sheet(self, result_file: str):
        """
        결과 파일의 규칙 시트에서 규칙 로드

        Args:
            result_file: 배정 결과 Excel 파일 경로
        """
        def is_empty(value):
            """빈 셀 체크 (NaN과 빈 문자열 모두 처리)"""
            return pd.isna(value) or (isinstance(value, str) and value.strip() == '')

        # header=None으로 읽어서 첫 번째 행도 데이터로 취급
        df = pd.read_excel(result_file, sheet_name='규칙', header=None)

        # 분반 규칙 파싱 (columns 1, 4)
        separation_count = 0
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 스킵
                continue

            student1_name = row.iloc[1]  # Column B (index 1)
            student2_name = row.iloc[4]  # Column E (index 4)

            # 빈 문자열도 체크하고 공백 제거
            if not is_empty(student1_name) and not is_empty(student2_name):
                name1 = str(student1_name).strip()
                name2 = str(student2_name).strip()
                self.separation_rules[name1].add(name2)
                self.separation_rules[name2].add(name1)
                self.separation_pairs.append((name1, name2))
                separation_count += 1

        # 합반 규칙 파싱 (columns 7, 10)
        together_count = 0
        current_group = set()
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 스킵
                continue

            student1_name = row.iloc[7]   # Column H (index 7)
            student2_name = row.iloc[10]  # Column K (index 10)

            # 빈 문자열도 빈 셀로 간주
            if not is_empty(student1_name) or not is_empty(student2_name):
                if not is_empty(student1_name):
                    current_group.add(str(student1_name).strip())
                if not is_empty(student2_name):
                    current_group.add(str(student2_name).strip())
            else:
                # 빈 행 = 그룹 종료
                if current_group:
                    self.together_groups.append(current_group)
                    together_count += len(current_group)
                    current_group = set()

        # 마지막 그룹 처리
        if current_group:
            self.together_groups.append(current_group)
            together_count += len(current_group)

        print(f"   ✅ 분반 규칙: {separation_count}쌍")
        print(f"   ✅ 합반 규칙: {len(self.together_groups)}그룹 ({together_count}명)")

        # 규칙 검증
        self._validate_rules()

    def _find_student_by_name(self, name: str) -> Optional[Student]:
        """이름으로 학생 찾기"""
        for student in self.students:
            if student.이름 == name:
                return student
        return None

    def _get_effective_count(self, class_num: int) -> int:
        """반의 유효 인원 계산 (특수반=3명, 전출생=0명, 일반=1명)"""
        return sum(s.effective_count() for s in self.classes[class_num])

    def _get_effective_gender_count(self, class_num: int, gender: str) -> int:
        """반의 특정 성별 유효 인원 계산"""
        return sum(s.effective_count() for s in self.classes[class_num]
                   if s.성별 == gender)

    def _can_assign(self, student: Student, class_num: int) -> bool:
        """학생을 특정 반에 배정할 수 있는지 검사 (분반 규칙 체크)"""
        # 이미 해당 반에 있는 학생들의 이름 목록
        students_in_class = [s.이름 for s in self.classes[class_num]]

        # 분반 규칙 확인
        names_to_avoid = self.separation_rules.get(student.이름, set())
        for name in students_in_class:
            if name in names_to_avoid:
                return False

        return True

    def _assign_student(self, student: Student, class_num: int, lock: bool = False):
        """학생을 특정 반에 배정"""
        if student.assigned_class is not None:
            # 이미 배정된 경우
            return

        # 분반 규칙 검증
        if not self._can_assign(student, class_num):
            # 배정할 수 없는 경우 - 다른 반 찾기
            for alternative_class in range(1, self.target_class_count + 1):
                if alternative_class != class_num and self._can_assign(student, alternative_class):
                    class_num = alternative_class
                    break
            else:
                # 어느 반에도 배정할 수 없음 - 오류
                print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (분반 규칙 충돌)")
                return

        student.assigned_class = class_num
        self.classes[class_num].append(student)

        if lock:
            student.locked = True

    def phase1_apply_rules(self):
        """Phase 1: 분반/합반 규칙 적용"""
        print("\n🎯 Phase 1: 분반/합반 규칙 적용 중...")

        # 먼저 합반 그룹 배정 (제약이 더 강함)
        for group_idx, group in enumerate(self.together_groups):
            # 그룹의 모든 학생 찾기
            group_students = []
            for name in group:
                student = self._find_student_by_name(name)
                if student:
                    group_students.append(student)
                else:
                    print(f"   ⚠️  경고: '{name}' 학생을 명단에서 찾을 수 없습니다.")

            if group_students:
                # 학생 수가 가장 적은 반에 배정
                target_class = min(self.classes.keys(),
                                 key=lambda c: len(self.classes[c]))

                for student in group_students:
                    self._assign_student(student, target_class, lock=True)

                print(f"   ✅ 합반 그룹 {group_idx + 1}: {[s.이름 for s in group_students]} → {target_class}반")

        # 분반 규칙 적용 (이미 배정된 학생들 고려)
        separation_applied = 0
        for name1, names_to_avoid in self.separation_rules.items():
            student1 = self._find_student_by_name(name1)
            if not student1:
                continue

            # student1이 미배정이면 먼저 배정
            if student1.assigned_class is None:
                # 유효 인원이 적은 반 우선, 같으면 해당 성별이 적은 반 우선
                target_class = min(range(1, self.target_class_count + 1),
                                  key=lambda c: (self._get_effective_count(c),
                                                self._get_effective_gender_count(c, student1.성별)))
                self._assign_student(student1, target_class, lock=True)
                separation_applied += 1

            # student1과 분반 규칙이 있는 학생들을 다른 반에 배정
            for name2 in names_to_avoid:
                student2 = self._find_student_by_name(name2)
                if student2 and student2.assigned_class is None:
                    # student1과 다른 반 중 유효 인원이 가장 적은 반 선택
                    # 유효 인원이 같으면 해당 성별이 적은 반 우선
                    available_classes = [c for c in range(1, self.target_class_count + 1) if c != student1.assigned_class]
                    target_class = min(available_classes,
                                     key=lambda c: (self._get_effective_count(c),
                                                   self._get_effective_gender_count(c, student2.성별)))
                    self._assign_student(student2, target_class, lock=True)
                    separation_applied += 1

        assigned_count = sum(1 for s in self.students if s.assigned_class is not None)
        print(f"   ✅ Phase 1 완료: {assigned_count}명 배정됨")

    def phase2_distribute_special_needs(self):
        """Phase 2: 특수반 학생 균등 배치"""
        print("\n🎯 Phase 2: 특수반 학생 균등 배치 중...")

        # 특수반 학생 현황 파악
        special_students = [s for s in self.students if s.특수반]
        assigned_special = [s for s in special_students if s.assigned_class is not None]
        unassigned_special = [s for s in special_students if s.assigned_class is None]

        print(f"   - 총 특수반 학생: {len(special_students)}명")
        print(f"   - 이미 배정됨: {len(assigned_special)}명")
        print(f"   - 배정 필요: {len(unassigned_special)}명")

        # 각 반의 현재 특수반 학생 수
        special_count_per_class = {c: sum(1 for s in self.classes[c] if s.특수반)
                                  for c in range(1, self.target_class_count + 1)}

        # 특수반 학생을 유효 인원이 적은 반부터 배정 (분반 규칙 고려)
        for student in unassigned_special:
            # 배정 가능한 반 중 유효 인원이 가장 적은 반 선택
            # (동점인 경우 특수반 학생이 적은 반 우선)
            valid_classes = [c for c in range(1, self.target_class_count + 1) if self._can_assign(student, c)]
            if valid_classes:
                target_class = min(valid_classes,
                                 key=lambda c: (self._get_effective_count(c), special_count_per_class[c]))
                self._assign_student(student, target_class, lock=True)
                special_count_per_class[target_class] += 1
            else:
                print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (규칙 충돌)")

        print(f"   ✅ 반별 특수반 학생 수: {special_count_per_class}")

    def phase3_separate_same_names(self):
        """Phase 3: 동명이인 분리"""
        print("\n🎯 Phase 3: 동명이인 분리 중...")

        # 이름별 빈도 계산
        name_counts = Counter(s.이름 for s in self.students)
        duplicate_names = {name: count for name, count in name_counts.items() if count > 1}

        if not duplicate_names:
            print("   ✅ 동명이인 없음")
            return

        print(f"   - 동명이인: {duplicate_names}")

        for name, count in duplicate_names.items():
            students_with_name = [s for s in self.students if s.이름 == name]
            assigned = [s for s in students_with_name if s.assigned_class is not None]
            unassigned = [s for s in students_with_name if s.assigned_class is None]

            # 이미 배정된 반 목록
            used_classes = {s.assigned_class for s in assigned}

            # 배정되지 않은 학생들을 다른 반에 배정
            for student in unassigned:
                # 동명이인이 없고 배정 가능한 반 중 유효 인원이 가장 적은 반 선택
                valid_classes = [c for c in range(1, self.target_class_count + 1)
                               if c not in used_classes and self._can_assign(student, c)]

                if valid_classes:
                    target_class = min(valid_classes,
                                     key=lambda c: self._get_effective_count(c))
                    self._assign_student(student, target_class, lock=True)
                    used_classes.add(target_class)
                else:
                    print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (동명이인/규칙 충돌)")

        print("   ✅ 동명이인 분리 완료")

    def phase4_balance_difficulty(self):
        """Phase 4: 난이도 균등 배분"""
        print("\n🎯 Phase 4: 난이도 균등 배분 중...")

        # 난이도가 있는 미배정 학생들
        unassigned = [s for s in self.students
                     if s.assigned_class is None and s.난이도 > 0]

        if not unassigned:
            print("   ✅ 배정할 난이도 학생 없음")
            return

        print(f"   - 난이도 배정 대상: {len(unassigned)}명")

        # 난이도가 높은 학생부터 배정 (균등 배분을 위해)
        unassigned.sort(key=lambda s: s.난이도, reverse=True)

        # 각 반의 현재 난이도 합
        difficulty_sum = {c: sum(s.난이도 for s in self.classes[c])
                         for c in range(1, self.target_class_count + 1)}

        for student in unassigned:
            # 배정 가능한 반 중 현재 난이도 합이 가장 낮은 반에 배정
            valid_classes = [c for c in range(1, self.target_class_count + 1) if self._can_assign(student, c)]
            if valid_classes:
                target_class = min(valid_classes,
                                 key=lambda c: difficulty_sum[c])
                self._assign_student(student, target_class, lock=True)
                difficulty_sum[target_class] += student.난이도
            else:
                print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (규칙 충돌)")

        print(f"   ✅ 반별 난이도 합: {difficulty_sum}")

    def phase5_balance_remaining(self):
        """Phase 5: 반별 순환 배정 (남녀 교차)"""
        print("\n🎯 Phase 5: 반별 순환 배정 중...")

        # 미배정 학생들
        unassigned = [s for s in self.students if s.assigned_class is None]

        if not unassigned:
            print("   ✅ 모든 학생 배정 완료")
            return

        print(f"   - 배정 대상: {len(unassigned)}명")

        # 1. 기존 반 처리 순서 랜덤 생성 (원본 반 수는 알 수 없으므로 unique 값 추출)
        original_classes = sorted(list(set(s.원반 for s in self.students)))
        random.shuffle(original_classes)
        print(f"   - 기존 반 처리 순서: {original_classes}")

        # 2. 각 기존 반별로 남녀 교차 처리
        for original_class in original_classes:
            # 2-1. 해당 반의 남학생 배정
            males = [s for s in self.students
                    if s.원반 == original_class and s.성별 == '남'
                    and s.assigned_class is None]
            males.sort(key=lambda s: s.점수, reverse=True)

            # 남학생 배정 시작 시점에 유효 인원 기준으로 반 정렬
            # 유효 인원이 같으면 남학생이 적은 반 우선
            target_classes = sorted(range(1, self.target_class_count + 1),
                                   key=lambda c: (self._get_effective_count(c),
                                                 self._get_effective_gender_count(c, '남')))

            for i, student in enumerate(males):
                target_class = target_classes[i % self.target_class_count]
                if self._can_assign(student, target_class):
                    self._assign_student(student, target_class, lock=False)
                else:
                    # 규칙 충돌 시 다음 반들 순서대로 시도
                    for offset in range(1, self.target_class_count):
                        alt_class = target_classes[(i + offset) % self.target_class_count]
                        if self._can_assign(student, alt_class):
                            self._assign_student(student, alt_class, lock=False)
                            break
                    else:
                        print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (규칙 충돌)")

            # 2-2. 해당 반의 여학생 배정
            females = [s for s in self.students
                      if s.원반 == original_class and s.성별 == '여'
                      and s.assigned_class is None]
            females.sort(key=lambda s: s.점수, reverse=True)

            # 여학생 배정 시작 시점에 다시 유효 인원 기준으로 반 정렬
            # 유효 인원이 같으면 여학생이 적은 반 우선
            target_classes = sorted(range(1, self.target_class_count + 1),
                                   key=lambda c: (self._get_effective_count(c),
                                                 self._get_effective_gender_count(c, '여')))

            for i, student in enumerate(females):
                target_class = target_classes[i % self.target_class_count]
                if self._can_assign(student, target_class):
                    self._assign_student(student, target_class, lock=False)
                else:
                    # 규칙 충돌 시 다음 반들 순서대로 시도
                    for offset in range(1, self.target_class_count):
                        alt_class = target_classes[(i + offset) % self.target_class_count]
                        if self._can_assign(student, alt_class):
                            self._assign_student(student, alt_class, lock=False)
                            break
                    else:
                        print(f"   ⚠️  경고: {student.이름} 학생을 배정할 수 없습니다 (규칙 충돌)")

        print("   ✅ 반별 순환 배정 완료")

    def phase6_random_distribution(self):
        """Phase 6: 랜덤 순환 배정"""
        print("\n🎯 Phase 6: 랜덤 순환 배정 중...")

        # 미배정 학생들 (Phase 5에서 모두 배정되어야 하지만 혹시 남은 경우 처리)
        unassigned = [s for s in self.students if s.assigned_class is None]

        if not unassigned:
            print("   ✅ 배정할 학생 없음 (모두 완료)")
            return

        print(f"   - 배정 대상: {len(unassigned)}명")

        # 성별로 분리 및 점수순 정렬
        males = sorted([s for s in unassigned if s.성별 == '남'],
                      key=lambda s: s.점수, reverse=True)
        females = sorted([s for s in unassigned if s.성별 == '여'],
                        key=lambda s: s.점수, reverse=True)

        # 랜덤 반 순서 생성
        male_order = list(range(1, 8))
        female_order = list(range(1, 8))
        random.shuffle(male_order)
        random.shuffle(female_order)

        print(f"   - 남학생 배정 순서: {male_order}")
        print(f"   - 여학생 배정 순서: {female_order}")

        # 남학생 순환 배정
        for idx, student in enumerate(males):
            class_num = male_order[idx % 7]
            self._assign_student(student, class_num, lock=False)

        # 여학생 순환 배정
        for idx, student in enumerate(females):
            class_num = female_order[idx % 7]
            self._assign_student(student, class_num, lock=False)

        print("   ✅ 랜덤 순환 배정 완료")

    def generate_output(self, output_file: str):
        """결과를 엑셀 파일로 출력"""
        print("\n📊 결과 생성 중...")

        # 스타일 정의
        THIN_BORDER = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        HEADER_FONT = Font(bold=True)
        HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # 연한 회색
        CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

        # 색상 팔레트 (30개 - 분반 쌍별 구분, 부족하면 재사용)
        COLOR_PALETTE = [
            "FFFF99", "FFCC99", "CCFFCC", "FFCCFF", "E6CCFF",
            "FFE6CC", "E6E6E6", "FFE6E6", "E6FFE6", "FFE6F0",
            "F0E6FF", "E6F0FF", "FFF0E6", "F0FFE6", "FFE6FF",
            "E6FFFF", "FFFFDD", "FFDDDD", "DDFFDD", "DDDDFF",
            "FFDDFF", "DDFFFF", "FFFFEE", "FFEEFF", "EEFFEE",
            "EEEEFF", "FFEEDD", "DDEEFF", "EEFFDD", "FFDDEE",
        ]

        # 합반 규칙 색상
        TOGETHER_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")  # 연한 파란색

        # 합반 규칙 학생 이름 집합 생성
        together_students = set()
        for group in self.together_groups:
            together_students.update(group)

        # 분반 쌍별 색상 매핑 생성 (여러 쌍에 속한 학생은 리스트로 저장)
        student_to_color = {}
        # 분반 상대방 정보 저장 (메모용)
        student_to_targets = defaultdict(set)
        
        for idx, (student1, student2) in enumerate(self.separation_pairs):
            color = COLOR_PALETTE[idx % len(COLOR_PALETTE)]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 각 학생이 속한 모든 쌍의 색상을 리스트로 저장
            if student1 not in student_to_color:
                student_to_color[student1] = []
            student_to_color[student1].append(fill)
            student_to_targets[student1].add(student2)

            if student2 not in student_to_color:
                student_to_color[student2] = []
            student_to_color[student2].append(fill)
            student_to_targets[student2].add(student1)

        # 디버깅: 색상 적용 대상 출력
        print(f"   📌 합반 규칙 학생: {sorted(together_students)}")
        print(f"   📌 분반 규칙: {len(self.separation_pairs)}쌍")
        print(f"   📌 분반 규칙 학생: {len(student_to_color)}명")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 시트 제거

        summary_data = []

        summary_data = []

        # 기준 학년 설정 (학생 데이터에서 가져옴, 없으면 기본값 5)
        base_grade = 5
        if self.students:
            # 가장 많은 학년을 기준으로 설정
            grades = [s.학년 for s in self.students]
            base_grade = max(set(grades), key=grades.count)
        
        target_grade = base_grade + 1

        # 각 반별 시트 생성
        for class_num in range(1, self.target_class_count + 1):
            students = self.classes[class_num]

            # 이름 가나다순 정렬
            students.sort(key=lambda s: s.이름)

            # 데이터프레임 생성
            data = []
            for idx, s in enumerate(students, 1):
                data.append({
                    '학년': target_grade,
                    '반': class_num,
                    '번호': idx,  # 새 반에서의 번호
                    '이름': s.이름,
                    '성별': s.성별,
                    '점수': s.점수,
                    '특수반': 1 if s.특수반 else '',
                    '전출': 1 if s.전출 else '',
                    '난이도': s.난이도 if s.난이도 > 0 else '',
                    '비고': s.비고,
                    '원학년': s.학년,
                    '원반': s.원반,
                    '원번호': s.원번호  # 원래 번호 유지
                })

            df = pd.DataFrame(data)

            # 시트 추가
            ws = wb.create_sheet(title=f'{target_grade}-{class_num}')
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 스타일 적용 (헤더)
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

            # 데이터 행 스타일 적용
            for row_idx, student in enumerate(students, start=2):  # 2부터 시작 (1은 헤더)
                # 기본 스타일: 테두리 및 정렬
                for col_idx in range(1, 14): # 13개 열
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    
                    # 가운데 정렬이 필요한 컬럼
                    # 학년(1), 반(2), 번호(3), 이름(4), 성별(5), 특수반(7), 전출(8), 원학년(11), 원반(12), 원번호(13)
                    if col_idx in [1, 2, 3, 4, 5, 7, 8, 11, 12, 13]:
                        cell.alignment = CENTER_ALIGN

                # 색상 및 메모 적용
                # 합반 규칙 학생인지 확인
                if student.이름 in together_students:
                    # 합반: 모든 셀에 동일한 파란색
                    for col_idx in range(1, 14):
                        ws.cell(row=row_idx, column=col_idx).fill = TOGETHER_FILL

                # 분반 규칙 학생인지 확인 (쌍별 색상)
                elif student.이름 in student_to_color:
                    fills = student_to_color[student.이름]
                    
                    # 메모 추가 (이름 셀인 4번 컬럼에)
                    targets = student_to_targets[student.이름]
                    target_info_list = []
                    for target_name in targets:
                        target_student = self._find_student_by_name(target_name)
                        if target_student and target_student.assigned_class:
                            target_info_list.append(f"{target_name}({target_student.assigned_class}반)")
                        else:
                            target_info_list.append(f"{target_name}(미배정)")
                    
                    target_str = ", ".join(sorted(target_info_list))
                    name_cell = ws.cell(row=row_idx, column=4)
                    name_cell.comment = Comment(f"분반 대상: {target_str}", "AutoAssigner")

                    if len(fills) == 1:
                        # 1개 쌍: 모든 셀에 동일한 색상
                        for col_idx in range(1, 14):
                            ws.cell(row=row_idx, column=col_idx).fill = fills[0]
                    else:
                        # 2개 이상 쌍: 홀수 셀과 짝수 셀에 번갈아가며 색상 적용
                        for col_idx in range(1, 14):
                            # 홀수 셀: 첫 번째 색상, 짝수 셀: 두 번째 색상
                            fill_idx = (col_idx - 1) % len(fills)
                            ws.cell(row=row_idx, column=col_idx).fill = fills[fill_idx]
                            
            # 컬럼 너비 자동 조정 (간단하게 고정값 적용)
            ws.column_dimensions['D'].width = 12 # 이름
            ws.column_dimensions['J'].width = 20 # 비고

            # 요약 데이터 수집
            summary_data.append({
                '반': f'{target_grade}-{class_num}',
                '학생수': len(students),
                '유효인원': sum(s.effective_count() for s in students),
                '남학생수': sum(1 for s in students if s.성별 == '남'),
                '여학생수': sum(1 for s in students if s.성별 == '여'),
                '유효남학생': sum(s.effective_count() for s in students if s.성별 == '남'),
                '유효여학생': sum(s.effective_count() for s in students if s.성별 == '여'),
                '난이도합': sum(s.난이도 for s in students),
                '특수반수': sum(1 for s in students if s.특수반),
                '전출생수': sum(1 for s in students if s.전출)
            })

            print(f"   ✅ {class_num}반 시트 생성: {len(students)}명")

        # 요약 시트 생성
        summary_df = pd.DataFrame(summary_data)
        ws_summary = wb.create_sheet(title='요약', index=0)
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws_summary.append(r)
            
        # 요약 시트 스타일
        for cell in ws_summary[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        for row in ws_summary.iter_rows(min_row=2):
            for cell in row:
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
        
        ws_summary.column_dimensions['A'].width = 10
        
        # 범례(Legend) 추가
        legend_start_row = len(summary_data) + 5
        ws_summary.cell(row=legend_start_row, column=1, value="[범례]").font = Font(bold=True, size=12)
        
        # 합반 범례
        ws_summary.cell(row=legend_start_row+1, column=1, value="합반 규칙 적용").border = THIN_BORDER
        guide_cell = ws_summary.cell(row=legend_start_row+1, column=2, value="하늘색 배경")
        guide_cell.fill = TOGETHER_FILL
        guide_cell.border = THIN_BORDER
        
        # 분반 범례
        ws_summary.cell(row=legend_start_row+2, column=1, value="분반 규칙 적용").border = THIN_BORDER
        guide_cell = ws_summary.cell(row=legend_start_row+2, column=2, value="기타 색상 배경")
        guide_cell.fill = PatternFill(start_color=COLOR_PALETTE[0], end_color=COLOR_PALETTE[0], fill_type="solid")
        guide_cell.border = THIN_BORDER
        guide_cell.comment = Comment("마우스를 올리면 누구와 분반인지 표시됩니다.", "AutoAssigner")
        
        ws_summary.cell(row=legend_start_row+2, column=3, value="← 이름에 마우스를 올리면 대상 확인 가능")

        # ==================== 규칙 시트 생성 ====================
        rules_ws = wb.create_sheet(title='규칙')

        # 헤더 작성
        headers = ['분반해야하는 학생', '', '', '', '', '', '', '합반해야 하는 학생', '', '', '']
        rules_ws.append(headers)

        # 합반 규칙을 행 리스트로 변환 (빈 행으로 그룹 구분)
        together_rows = []
        for group in self.together_groups:
            group_list = sorted(list(group))  # 정렬하여 일관성 유지
            # 2명씩 쌍으로 묶기
            for i in range(0, len(group_list), 2):
                name1 = group_list[i]
                name2 = group_list[i + 1] if i + 1 < len(group_list) else None
                together_rows.append((name1, name2))
            # 그룹 구분용 빈 행
            together_rows.append((None, None))

        # 마지막 빈 행 제거
        if together_rows and together_rows[-1] == (None, None):
            together_rows.pop()

        # 데이터 추가
        max_rows = max(len(self.separation_pairs), len(together_rows) if together_rows else 0)

        for i in range(max_rows):
            row = [None] * 11  # 11개 컬럼 (A-K), None으로 초기화

            # 분반 규칙 (columns 1, 4 → index 1, 4)
            if i < len(self.separation_pairs):
                name1, name2 = self.separation_pairs[i]
                row[1] = name1
                row[4] = name2

            # 합반 규칙 (columns 7, 10 → index 7, 10)
            if i < len(together_rows):
                name1, name2 = together_rows[i]
                row[7] = name1 if name1 else None
                row[10] = name2 if name2 else None

            rules_ws.append(row)

        # 스타일링 (헤더)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_font = Font(bold=True)
        for cell in rules_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"   ✅ 규칙 시트 추가: 분반 {len(self.separation_pairs)}쌍, 합반 {len(self.together_groups)}그룹")

        # 파일 저장
        wb.save(output_file)
        print(f"\n✅ 결과 파일 저장: {output_file}")

        # 요약 출력
        print("\n" + "=" * 70)
        print("📋 반별 요약")
        print("=" * 70)
        print(summary_df.to_string(index=False))

    def run(self, output_file: str = "03 6학년 배정 결과.xlsx"):
        """전체 프로세스 실행"""
        try:
            # 데이터 로드
            self.load_students()
            self.load_rules()

            # 6단계 배정 프로세스
            self.phase1_apply_rules()
            self.phase2_distribute_special_needs()
            self.phase3_separate_same_names()
            self.phase4_balance_difficulty()
            self.phase5_balance_remaining()
            # self.phase6_random_distribution()  # Phase 5에서 모두 처리하므로 비활성화

            # 결과 생성
            self.generate_output(output_file)

            print("\n" + "=" * 70)
            print("🎉 학급 편성 완료!")
            print("=" * 70)

        except Exception as e:
            print(f"\n❌ 오류 발생: {e}")
            raise


def get_base_path():
    """실행 파일의 경로를 반환 (PyInstaller 지원)"""
    if getattr(sys, 'frozen', False):
        # PyInstaller로 빌드된 실행 파일
        return os.path.dirname(sys.executable)
    else:
        # 일반 Python 스크립트
        return os.path.dirname(os.path.abspath(__file__))


def select_file(title, filetypes, default_path=None, mode='open'):
    """
    파일 선택 다이얼로그

    Args:
        title: 다이얼로그 제목
        filetypes: 파일 타입 리스트 [("Excel files", "*.xlsx"), ...]
        default_path: 기본 파일 경로
        mode: 'open' (파일 열기) 또는 'save' (파일 저장)

    Returns:
        선택된 파일 경로 (취소시 None)
    """
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기

    # 기본 경로 설정
    if default_path and os.path.exists(os.path.dirname(default_path)):
        initialdir = os.path.dirname(default_path)
        initialfile = os.path.basename(default_path)
    else:
        initialdir = get_base_path()
        initialfile = ""

    # 파일 선택 다이얼로그
    if mode == 'open':
        file_path = filedialog.askopenfilename(
            title=title,
            filetypes=filetypes,
            initialdir=initialdir,
            initialfile=initialfile if os.path.exists(default_path or "") else ""
        )
    else:  # save
        file_path = filedialog.asksaveasfilename(
            title=title,
            filetypes=filetypes,
            initialdir=initialdir,
            initialfile=initialfile,
            defaultextension=".xlsx"
        )

    root.destroy()
    return file_path if file_path else None


def main():
    """메인 함수"""
    print("=" * 70)
    print("🎓 자동 학급 편성 프로그램")
    print("=" * 70)

    base_path = get_base_path()
    excel_filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]

    # 기본 파일 경로
    default_student_file = os.path.join(base_path, '01 5학년_가상 명단.xlsx')
    default_rules_file = os.path.join(base_path, '02 분반 합반할 학생 규칙.xlsx')
    default_output_file = os.path.join(base_path, '03 6학년 배정 결과.xlsx')

    # 1. 학생 명단 파일 선택
    print("\n📂 Step 1: 5학년 학생 명단 파일을 선택하세요...")
    student_file = select_file(
        title="5학년 학생 명단 파일을 선택하세요",
        filetypes=excel_filetypes,
        default_path=default_student_file,
        mode='open'
    )

    if not student_file:
        print("❌ 파일 선택이 취소되었습니다.")
        if sys.stdin.isatty():
            input("\n종료하려면 Enter를 누르세요...")
        sys.exit(0)

    print(f"   ✅ 선택됨: {os.path.basename(student_file)}")

    # 2. 규칙 파일 선택
    print("\n📂 Step 2: 분반/합반 규칙 파일을 선택하세요...")
    rules_file = select_file(
        title="분반/합반 규칙 파일을 선택하세요",
        filetypes=excel_filetypes,
        default_path=default_rules_file,
        mode='open'
    )

    if not rules_file:
        print("❌ 파일 선택이 취소되었습니다.")
        if sys.stdin.isatty():
            input("\n종료하려면 Enter를 누르세요...")
        sys.exit(0)

    print(f"   ✅ 선택됨: {os.path.basename(rules_file)}")

    # 3. 출력 파일 위치 선택
    print("\n📂 Step 3: 결과 파일을 저장할 위치를 선택하세요...")
    output_file = select_file(
        title="결과 파일 저장 위치",
        filetypes=excel_filetypes,
        default_path=default_output_file,
        mode='save'
    )

    if not output_file:
        print("❌ 파일 선택이 취소되었습니다.")
        if sys.stdin.isatty():
            input("\n종료하려면 Enter를 누르세요...")
        sys.exit(0)

    print(f"   ✅ 저장 위치: {os.path.basename(output_file)}")

    # 4. 학급 편성 실행
    try:
        assigner = ClassAssigner(
            student_file=student_file,
            rules_file=rules_file
        )
        assigner.run(output_file=output_file)

        print("\n" + "=" * 70)
        print(f"✅ 결과 파일이 생성되었습니다!")
        print(f"📁 위치: {output_file}")
        print("=" * 70)

        # 완료 메시지 박스 (GUI)
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "완료",
            f"학급 편성이 완료되었습니다!\n\n결과 파일:\n{output_file}"
        )
        root.destroy()

    except Exception as e:
        print(f"\n❌ 오류가 발생했습니다: {e}")
        import traceback
        traceback.print_exc()

        # 오류 메시지 박스 (GUI)
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "오류",
            f"오류가 발생했습니다:\n\n{str(e)}\n\n자세한 내용은 콘솔을 확인하세요."
        )
        root.destroy()

        if sys.stdin.isatty():
            input("\n종료하려면 Enter를 누르세요...")
        sys.exit(1)


if __name__ == '__main__':
    main()
